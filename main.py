#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Farol ONA - NiceGUI >= 3.0 — v5.5 (fix: gráfico NC com formatação correta)
"""
from __future__ import annotations
import os, re, io, csv, zipfile, shutil, ast as _ast, threading, difflib, json, hashlib, secrets, asyncio
import logging
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any

from nicegui import ui, app

ui.add_head_html("""<style>
.farol-modal-card { 
  border-radius: 1rem !important; 
  display: flex !important; 
  flex-direction: column !important; 
  max-height: 95vh !important;
  overflow-y: auto !important;
  overflow-x: hidden !important;
}

/* Desktop: largura total com máximo */
@media (min-width: 1024px) {
  .farol-modal-card {
    width: 980px !important;
    max-height: 90vh !important;
  }
}

/* Tablet: reduz um pouco */
@media (max-width: 1023px) and (min-width: 640px) {
  .farol-modal-card {
    width: 95vw !important;
    max-height: 92vh !important;
  }
}

/* Mobile: ocupa a maior parte possível */
@media (max-width: 639px) {
  .farol-modal-card {
    width: 98vw !important;
    max-height: 98vh !important;
  }
}

.farol-modal-close { position: absolute !important; top: 1rem; right: 1rem; }
.farol-textarea .q-field__control { min-height: 120px !important; overflow-y: auto !important; }
.farol-modal-content { 
  flex: 1 !important;
  overflow-y: auto !important;
}
</style>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="theme-color" content="#10b981">
""", shared=True)


ui.add_head_html("""<style>
.farol-mobile-only { display: flex; }
.farol-desktop-only { display: none; }
.farol-mobile-drawer { display: block; }
@media (min-width: 768px) {
  .farol-mobile-only { display: none !important; }
  .farol-desktop-only { display: flex !important; }
  .farol-mobile-drawer { display: none !important; }
}
</style>""", shared=True)

from openpyxl import load_workbook

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DATA_DIR     = os.path.join(BASE_DIR, "data")
DEFAULT_XLSX = os.path.join(DATA_DIR, "Farol de acompanhamento ONA I BASE.xlsx")
USERS_JSON   = os.path.join(DATA_DIR, "users.json")
DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "6105/*"


AVALIACAO_OPCOES = ["Conforme","Parcialmente Conforme","Não Conforme","Não Avaliado"]

H_SECAO     = "Seção"
H_REQ_GERAL = "Requisito geral"
H_REQ       = "Requisitos"
H_AVAL      = "Avaliação"
H_EVID      = "Evidências"
H_OBS       = "Observações"
H_AUDIT_DATE = "Data Alteração"
H_AUDIT_USER = "Usuário Alteração"


HEADER_ALIASES: Dict[str, List[str]] = {
    "secao":           ["seção","secao","section"],
    "requisito geral": ["requisito geral","requisito_geral","req geral","requisito global"],
    "requisitos":      ["requisitos","requisito","req","itens",
                        "requisitos específicos","requisitos especificos"],
    "avaliacao":       ["avaliação","avaliacao","status"],
    "evidencias":      ["evidências","evidencias","evidência","evidencia"],
    "observacoes":     ["observações","observacoes","observação","observacao","obs"],
    "prazo":           ["prazo","prazo de resolução","prazo de resolucao","data de conclusão","data conclusão","deadline","vencimento"],
    "data alteracao":  ["data alteração","data alteracao","última alteração","ultima alteracao","modificado em","alterado em"],
    "usuario alteracao":["usuário alteração","usuario alteracao","alterado por","modificado por","usuário","usuario"],
}
REQUIRED_BASE = {"requisito geral","requisitos","avaliacao"}
OPTIONAL_PREFERRED = {"evidencias","observacoes","prazo"}
INITIAL_RENDER_LIMIT = 120
RENDER_STEP          = 120

def safe_mkdir(p): os.makedirs(p, exist_ok=True)
def is_blank(v): return v is None or (isinstance(v, str) and v.strip() == "")

def norm(s):
    s = str(s or "").strip().lower()
    for a,b in [("ã","a"),("á","a"),("à","a"),("â","a"),("é","e"),("ê","e"),
                ("è","e"),("í","i"),("ó","o"),("ô","o"),("õ","o"),("ú","u"),
                ("ç","c"),("\n"," ")]:
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s)

def slugify(s):
    s = norm(s)
    s = re.sub(r"\s+", "-", s)
    return re.sub(r"[^a-z0-9\-]", "", s)


def natural_sort_key(value: Any):
    parts = re.split(r'(\d+)', str(value or '').strip().lower())
    out = []
    for part in parts:
        if part.isdigit():
            out.append(int(part))
        else:
            out.append(part)
    return out


def _hash_password(password: str, salt: str) -> str:
    return hashlib.sha256(f"{salt}:{password}".encode("utf-8")).hexdigest()


def ensure_users_file() -> None:
    safe_mkdir(DATA_DIR)
    if os.path.exists(USERS_JSON):
        return
    salt = secrets.token_hex(16)
    payload = {
        "version": 1,
        "users": [
            {
                "username": DEFAULT_ADMIN_USERNAME,
                "password_hash": _hash_password(DEFAULT_ADMIN_PASSWORD, salt),
                "salt": salt,
                "is_admin": True,
                "created_at": datetime.now().isoformat(timespec="seconds"),
            }
        ],
    }
    with open(USERS_JSON, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def load_users() -> Dict[str, Any]:
    ensure_users_file()
    with open(USERS_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_users(data: Dict[str, Any]) -> None:
    safe_mkdir(DATA_DIR)
    with open(USERS_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_user_record(username: str) -> Optional[Dict[str, Any]]:
    username_n = (username or '').strip().lower()
    if not username_n:
        return None
    data = load_users()
    for user in data.get('users', []):
        if (user.get('username') or '').strip().lower() == username_n:
            return user
    return None


def verify_login(username: str, password: str) -> Optional[Dict[str, Any]]:
    user = get_user_record(username)
    if not user:
        return None
    salt = user.get('salt') or ''
    if user.get('password_hash') == _hash_password(password or '', salt):
        return user
    return None


def create_user(username: str, password: str, is_admin: bool = False) -> Tuple[bool, str]:
    username_clean = (username or '').strip()
    if len(username_clean) < 3:
        return False, 'Usuário deve ter pelo menos 3 caracteres.'
    if len(password or '') < 4:
        return False, 'Senha deve ter pelo menos 4 caracteres.'
    data = load_users()
    if any((u.get('username') or '').strip().lower() == username_clean.lower() for u in data.get('users', [])):
        return False, 'Usuário já existe.'
    salt = secrets.token_hex(16)
    data.setdefault('users', []).append({
        'username': username_clean,
        'password_hash': _hash_password(password, salt),
        'salt': salt,
        'is_admin': bool(is_admin),
        'created_at': datetime.now().isoformat(timespec='seconds'),
    })
    save_users(data)
    return True, 'Usuário criado com sucesso.'


def get_session() -> Dict[str, Any]:
    return app.storage.user


def current_username() -> str:
    return str(get_session().get('username') or '').strip()


def current_is_admin() -> bool:
    return bool(get_session().get('is_admin'))


def login_user(user: Dict[str, Any]) -> None:
    session = get_session()
    session['authenticated'] = True
    session['username'] = user.get('username')
    session['is_admin'] = bool(user.get('is_admin'))


def logout_user() -> None:
    session = get_session()
    session.clear()


def require_login() -> bool:
    if not get_session().get('authenticated'):
        ui.navigate.to('/login')
        return False
    return True


def require_admin() -> bool:
    if not require_login():
        return False
    if not current_is_admin():
        notify_err('Acesso restrito ao administrador.')
        ui.navigate.to('/')
        return False
    return True


def fuzzy_best_match(value: str, aliases_map: Dict[str, List[str]], threshold: float = 0.84) -> Optional[str]:
    value_n = norm(value)
    if not value_n:
        return None
    best_key = None
    best_score = 0.0
    for key, candidates in aliases_map.items():
        for candidate in candidates:
            score = difflib.SequenceMatcher(None, value_n, candidate).ratio()
            if score > best_score:
                best_score = score
                best_key = key
    return best_key if best_score >= threshold else None

def status_color(status):
    t = norm(status)
    if "parcial" in t:                            return "orange"
    if "nao conforme" in t or "nao conform" in t: return "red"
    if "conforme" in t and "parcial" not in t:    return "green"
    if "nao aval" in t:                           return "grey"
    return "grey"

def normalize_status(status):
    t = norm(status)
    if not t or t in ("nan","none",""):           return "Não Avaliado"
    if "parcial" in t:                            return "Parcialmente Conforme"
    if "nao conforme" in t or "nao conform" in t: return "Não Conforme"
    if "conforme" in t:                           return "Conforme"
    if "nao aval" in t:                           return "Não Avaliado"
    return str(status).strip() or "Não Avaliado"

def notify_warn(msg): ui.notify(msg, type="warning",  close_button="OK")
def notify_ok(msg):   ui.notify(msg, type="positive", close_button="OK")
def notify_err(msg):  ui.notify(msg, type="negative", close_button="OK")

def is_valid_xlsx_path(path):
    try:    return zipfile.is_zipfile(path)
    except: return False

def is_valid_xlsx_bytes(raw):
    try:    return zipfile.is_zipfile(io.BytesIO(raw))
    except: return False

def clamp01(v): return min(max(v, 0.0), 1.0)

# ── Navbar ────────────────────────────────────────────────────────────────────

def get_save_status_meta():
    if STATE.is_save_pending():
        return '💾✏️ Salvando...', 'amber'
    if STATE.last_save_error:
        return '💾⚠️ Falha ao salvar', 'red'
    if STATE.last_save_at:
        return f'💾✅ Salvo às {STATE.last_save_at[-8:]}', 'green'
    return '💾 Sem alterações pendentes', 'grey'



def render_navbar(active="home"):
    username = current_username() or "Usuário"

    def go(route: str):
        ui.navigate.to(route)

    def nav_items_drawer():
        with ui.column().classes("gap-2 w-full"):
            def item(label, icon, route, key):
                color = "primary" if key == active else "grey-8"
                ui.button(
                    label,
                    icon=icon,
                    on_click=lambda r=route: (mobile_drawer.toggle(), go(r)),
                ).props(f"flat no-caps align=left color={color}").classes("w-full justify-start")
            item("Home", "home", "/", "home")
            item("Dashboard", "dashboard", "/dashboard", "dashboard")
            if current_is_admin():
                item("Configurações", "settings", "/config", "config")

    with ui.left_drawer(value=False, top_corner=False, bottom_corner=True).classes(
        "farol-mobile-drawer bg-white border-r border-gray-200 w-[82vw] max-w-[320px]"
    ) as mobile_drawer:
        with ui.column().classes("w-full gap-4 p-4"):
            with ui.row().classes("items-center gap-3"):
                with ui.element("div").classes("w-10 h-10 flex items-center justify-center"):
                    ui.image("./data/logo.png").classes("w-full h-full object-contain")
                with ui.column().classes("gap-0"):
                    ui.label("Farol ONA").classes("text-lg font-bold text-emerald-700 leading-tight")
                    ui.label("Santa Casa").classes("text-xs opacity-70")
            ui.separator()
            nav_items_drawer()
            ui.separator()
            status_badge_m = ui.badge('', color='grey').classes('w-full justify-center text-center')
            def refresh_save_badge_m():
                label, color = get_save_status_meta()
                status_badge_m.set_text(label)
                status_badge_m.props(f'color={color}')
            refresh_save_badge_m()
            ui.timer(0.7, refresh_save_badge_m)
            with ui.row().classes("items-center gap-2 flex-wrap"):
                ui.badge(username, color="indigo")
                if current_is_admin():
                    ui.badge("Admin", color="deep-orange")
            ui.button(
                "Sair",
                icon="logout",
                on_click=lambda: (mobile_drawer.toggle(), logout_user(), ui.navigate.to('/login')),
            ).props("outline no-caps").classes("w-full")

    with ui.header().classes("bg-white text-gray-900 shadow-sm border-b border-gray-200"):
        with ui.row().classes("w-full items-center justify-between px-3 sm:px-4 md:px-6 py-2 md:py-3 max-w-[1400px] mx-auto gap-2"):
            with ui.row().classes("items-center gap-3 min-w-0 flex-1 cursor-pointer").on("click", lambda: go("/")):
                with ui.element("div").classes("w-9 h-9 md:w-12 md:h-12 flex items-center justify-center shrink-0"):
                    ui.image("./data/logo.png").classes("w-full h-full object-contain").style(
                        "filter: drop-shadow(0 2px 4px rgba(0,0,0,0.1))"
                    )
                with ui.column().classes("gap-0 min-w-0"):
                    ui.label("Farol ONA").classes("text-xl md:text-2xl font-bold text-emerald-700 leading-tight truncate")
                    ui.label("Santa Casa de Poços de Caldas").classes("hidden sm:block text-xs opacity-70 font-medium truncate")

            with ui.row().classes("farol-desktop-only items-center gap-1"):
                def nav_btn(label, icon, route, key):
                    props = "flat no-caps" + (" color=primary" if key == active else "")
                    ui.button(label, icon=icon, on_click=lambda r=route: go(r)).props(props).classes("px-3")
                nav_btn("Home", "home", "/", "home")
                nav_btn("Dashboard", "dashboard", "/dashboard", "dashboard")
                if current_is_admin():
                    nav_btn("Configurações", "settings", "/config", "config")

            with ui.row().classes("farol-desktop-only items-center gap-2 ml-4"):
                save_badge = ui.badge('', color='grey').classes('min-w-[170px] justify-center')
                def refresh_save_badge():
                    label, color = get_save_status_meta()
                    save_badge.set_text(label)
                    save_badge.props(f'color={color}')
                refresh_save_badge()
                ui.timer(0.7, refresh_save_badge)
                ui.badge(username, color="indigo")
                if current_is_admin():
                    ui.badge("Admin", color="deep-orange")
                ui.button(
                    "Sair",
                    icon="logout",
                    on_click=lambda: (logout_user(), ui.navigate.to('/login')),
                ).props("outline dense no-caps")

            with ui.row().classes("farol-mobile-only items-center gap-2 shrink-0"):
                save_badge_compact = ui.badge('', color='grey').classes('max-w-[145px] justify-center text-[11px]')
                def refresh_save_badge_compact():
                    label, color = get_save_status_meta()
                    compact = label.replace("Sem alterações pendentes", "Sem pendências").replace("Salvo às ", "Salvo ")
                    save_badge_compact.set_text(compact)
                    save_badge_compact.props(f'color={color}')
                refresh_save_badge_compact()
                ui.timer(0.7, refresh_save_badge_compact)
                ui.button(icon="menu", on_click=mobile_drawer.toggle).props("flat round color=primary")

def csv_bytes(rows, filename_hint="export.csv"):
    has_secao = any(r.get("secao","").strip() for r in rows)
    cols = ([H_SECAO] if has_secao else []) + [H_REQ_GERAL, H_REQ, H_AVAL, H_EVID, H_OBS]
    buf  = io.StringIO()
    w    = csv.DictWriter(buf, fieldnames=cols, delimiter=";", extrasaction="ignore")
    w.writeheader()
    for r in rows:
        row_out = {H_REQ_GERAL: r.get("requisito_geral",""), H_REQ: r.get("requisito",""),
                   H_AVAL: r.get("avaliacao",""), H_EVID: r.get("evidencias",""),
                   H_OBS: r.get("observacoes","")}
        if has_secao: row_out[H_SECAO] = r.get("secao","")
        w.writerow(row_out)
    return buf.getvalue().encode("utf-8-sig"), filename_hint

# ── Modelos ───────────────────────────────────────────────────────────────────
@dataclass
class SheetInfo:
    name: str; state: str; slug: str

@dataclass
class RowItem:
    secao: str; requisito_geral: str; requisito: str
    avaliacao: str; evidencias: str; observacoes: str; excel_row: int; prazo: str = ""
    data_alteracao: str = ""; usuario_alteracao: str = ""
    def to_dict(self):
        return {k: getattr(self,k) for k in
                ("secao","requisito_geral","requisito","avaliacao","evidencias","observacoes","prazo","data_alteracao","usuario_alteracao")}

@dataclass
class SectionStat:
    nome: str; total: int; avaliados: int
    conforme: int; parcial: int; nao_conforme: int; nav: int
    @property
    def pct_av(self):   return (self.avaliados / self.total * 100.0) if self.total else 0.0
    @property
    def pct_conf(self): return (self.conforme  / self.total * 100.0) if self.total else 0.0

# ── ExcelEngine ───────────────────────────────────────────────────────────────
class ExcelEngine:
    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path
        self.sheet_infos: List[SheetInfo]                    = []
        self.cache_items:    Dict[str, List[RowItem]]        = {}
        self.cache_secstats: Dict[str, List[SectionStat]]   = {}
        self.cache_header_map: Dict[str, Tuple]              = {}
        self.last_loaded: Optional[datetime]                 = None
        self._read_wb = None
        self._read_mtime: Optional[float] = None
        self._read_lock = threading.Lock()

    def exists(self): return os.path.exists(self.xlsx_path)

    def invalidate_read_cache(self):
        with self._read_lock:
            if self._read_wb is not None:
                try:
                    self._read_wb.close()
                except Exception:
                    pass
            self._read_wb = None
            self._read_mtime = None

    def _get_read_workbook(self):
        if not self.exists() or not is_valid_xlsx_path(self.xlsx_path):
            return None
        try:
            mtime = os.path.getmtime(self.xlsx_path)
        except OSError:
            return None
        with self._read_lock:
            if self._read_wb is not None and self._read_mtime == mtime:
                return self._read_wb
            if self._read_wb is not None:
                try:
                    self._read_wb.close()
                except Exception:
                    pass
            try:
                self._read_wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
                self._read_mtime = mtime
                return self._read_wb
            except Exception as ex:
                print(f"[WARN] workbook cache load: {ex}")
                self._read_wb = None
                self._read_mtime = None
                return None

    def list_sheets(self):
        if not self.exists() or not is_valid_xlsx_path(self.xlsx_path):
            self.sheet_infos = []; return []
        wb = self._get_read_workbook()
        if wb is None:
            self.sheet_infos = []; return []
        out = []
        for s in wb.sheetnames:
            ws = wb[s]
            state = getattr(ws,"sheet_state","visible") or "visible"
            out.append(SheetInfo(name=s, state=state, slug=slugify(s)))
        self.sheet_infos = out
        self.last_loaded = datetime.now()
        return out

    def visible_chapters(self):
        infos = self.sheet_infos or self.list_sheets()
        chapters = [si for si in infos if norm(si.name) != "dashboard" and si.state == "visible"]
        return sorted(chapters, key=lambda si: natural_sort_key(si.name))

    def _find_header(self, ws, max_rows=200):
        aliases = {k: [norm(x) for x in v] for k, v in HEADER_ALIASES.items()}
        best_candidate = None
        best_score = -1

        for r_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_i > max_rows:
                break
            if not row:
                continue

            found: Dict[str, int] = {}
            score = 0

            for c_i, v in enumerate(row, start=1):
                if is_blank(v):
                    continue
                cn = norm(v)

                exact_key = None
                for key, alist in aliases.items():
                    if key not in found and cn in alist:
                        exact_key = key
                        break
                if exact_key:
                    found[exact_key] = c_i
                    score += 3
                    continue

                fuzzy_key = fuzzy_best_match(cn, aliases, threshold=0.84)
                if fuzzy_key and fuzzy_key not in found:
                    found[fuzzy_key] = c_i
                    score += 1

            base_ok = REQUIRED_BASE.issubset(found.keys())
            useful_optional = len(OPTIONAL_PREFERRED.intersection(found.keys()))
            if base_ok:
                score += useful_optional * 2
                if "secao" in found:
                    score += 1
                if score > best_score:
                    best_score = score
                    best_candidate = (r_i, found)

        if not best_candidate:
            raise ValueError("Cabeçalho não encontrado nas primeiras 200 linhas.")

        r_i, found = best_candidate
        col_map = {
            "requisito_geral": found["requisito geral"],
            "requisito":       found["requisitos"],
            "avaliacao":       found["avaliacao"],
            "evidencias":      found.get("evidencias"),
            "observacoes":     found.get("observacoes"),
        }
        if "secao" in found:
            col_map["secao"] = found["secao"]
        if "prazo" in found:
            col_map["prazo"] = found["prazo"]
        if "data alteracao" in found:
            col_map["data_alteracao"] = found["data alteracao"]
        if "usuario alteracao" in found:
            col_map["usuario_alteracao"] = found["usuario alteracao"]
        return r_i, col_map

    def load_items(self, sheet_name, force=False):
        if not force and sheet_name in self.cache_items:
            return self.cache_items[sheet_name]
        if not self.exists() or not is_valid_xlsx_path(self.xlsx_path):
            self.cache_items[sheet_name] = []; return []
        wb = self._get_read_workbook()
        if wb is None:
            self.cache_items[sheet_name] = []; return []
        if sheet_name not in wb.sheetnames:
            self.cache_items[sheet_name] = []; return []
        ws = wb[sheet_name]
        header_row, col_map = self._find_header(ws)
        self.cache_header_map[sheet_name] = (header_row, col_map)
        has_secao = "secao" in col_map
        last_secao = ""; last_req_geral = ""
        items: List[RowItem] = []
        for r_i, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
            if not row or all(is_blank(x) for x in row):
                continue

            def getv(col_1b):
                if not col_1b:
                    return ""
                idx = col_1b - 1
                if idx < 0 or idx >= len(row):
                    return ""
                v = row[idx]
                return "" if v is None else str(v).strip()

            req_geral = getv(col_map["requisito_geral"])
            req       = getv(col_map["requisito"])
            aval      = getv(col_map["avaliacao"])
            evid      = getv(col_map.get("evidencias"))
            obs       = getv(col_map.get("observacoes"))
            prazo     = getv(col_map.get("prazo"))
            data_alt  = getv(col_map.get("data_alteracao"))
            user_alt  = getv(col_map.get("usuario_alteracao"))
            secao     = getv(col_map.get("secao")) if has_secao else ""
            if is_blank(req):
                continue
            if not is_blank(secao):
                last_secao = secao.strip()
            else:
                secao = last_secao
            if not is_blank(req_geral):
                last_req_geral = req_geral.strip()
            else:
                req_geral = last_req_geral
            items.append(RowItem(
                secao=secao.strip(),
                requisito_geral=(req_geral or "").strip() or "(Sem título)",
                requisito=(req or "").strip(),
                avaliacao=normalize_status(aval),
                evidencias=(evid or "").strip(),
                observacoes=(obs or "").strip(),
                excel_row=r_i,
                prazo=(prazo or "").strip(),
                data_alteracao=(data_alt or "").strip(),
                usuario_alteracao=(user_alt or "").strip(),
            ))
        self.cache_items[sheet_name] = items
        self.cache_secstats.pop(sheet_name, None)
        return items

    def section_stats(self, sheet_name):
        if sheet_name in self.cache_secstats:
            return self.cache_secstats[sheet_name]
        items = self.load_items(sheet_name)
        if not items:
            self.cache_secstats[sheet_name] = []; return []
        has_secao = any(it.secao.strip() for it in items)
        key_fn = ((lambda it: it.secao.strip() or "(Sem seção)")
                  if has_secao
                  else (lambda it: it.requisito_geral.strip() or "(Sem título)"))
        groups: Dict[str, List[RowItem]] = {}
        for it in items:
            groups.setdefault(key_fn(it), []).append(it)
        stats: List[SectionStat] = []
        for nome, grp in groups.items():
            total = len(grp)
            conf  = sum(1 for x in grp if x.avaliacao == "Conforme")
            parc  = sum(1 for x in grp if x.avaliacao == "Parcialmente Conforme")
            nconf = sum(1 for x in grp if x.avaliacao == "Não Conforme")
            nav   = sum(1 for x in grp if x.avaliacao == "Não Avaliado")
            stats.append(SectionStat(nome=nome, total=total, avaliados=conf+parc+nconf,
                                     conforme=conf, parcial=parc, nao_conforme=nconf, nav=nav))
        self.cache_secstats[sheet_name] = stats
        return stats

    def save_items(self, sheet_name, items, changed_by: str = ""):
        if not self.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {self.xlsx_path}")
        
        # Backup assíncrono em background (não bloqueia o salvamento)
        bak = self.xlsx_path + ".bak"
        if not os.path.exists(bak):
            shutil.copy2(self.xlsx_path, bak)
        
        wb = load_workbook(self.xlsx_path, data_only=False)
        try:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Aba '{sheet_name}' não existe.")
            ws = wb[sheet_name]
            header_row, col_map = self._find_header(ws)

            next_col = ws.max_column + 1
            if not col_map.get("data_alteracao"):
                col_map["data_alteracao"] = next_col
                ws.cell(row=header_row, column=col_map["data_alteracao"]).value = H_AUDIT_DATE
                next_col += 1
            if not col_map.get("usuario_alteracao"):
                col_map["usuario_alteracao"] = next_col
                ws.cell(row=header_row, column=col_map["usuario_alteracao"]).value = H_AUDIT_USER
                next_col += 1

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            actor = (changed_by or "sistema").strip()

            for it in items:
                r = it.excel_row
                if r <= header_row:
                    continue
                ws.cell(row=r, column=col_map["avaliacao"]).value = it.avaliacao
                if col_map.get("evidencias"):
                    ws.cell(row=r, column=col_map["evidencias"]).value = it.evidencias
                if col_map.get("observacoes"):
                    ws.cell(row=r, column=col_map["observacoes"]).value = it.observacoes
                if col_map.get("prazo"):
                    ws.cell(row=r, column=col_map["prazo"]).value = it.prazo
                ws.cell(row=r, column=col_map["data_alteracao"]).value = timestamp
                ws.cell(row=r, column=col_map["usuario_alteracao"]).value = actor
                it.data_alteracao = timestamp
                it.usuario_alteracao = actor
            wb.save(self.xlsx_path)
            self.invalidate_read_cache()
            self.cache_items.pop(sheet_name, None)
            self.cache_secstats.pop(sheet_name, None)
            self.cache_header_map.pop(sheet_name, None)
            self.sheet_infos = []
        finally: wb.close()

    def compute_dashboard(self):
        rows = []
        for si in self.visible_chapters():
            items = self.load_items(si.name)
            if not items: continue
            total     = len(items)
            conf      = sum(1 for x in items if x.avaliacao == "Conforme")
            parc      = sum(1 for x in items if x.avaliacao == "Parcialmente Conforme")
            nconf     = sum(1 for x in items if x.avaliacao == "Não Conforme")
            nav       = sum(1 for x in items if x.avaliacao == "Não Avaliado")
            avaliados = conf + parc + nconf
            rows.append({
                "Capítulo": si.name, "slug": si.slug, "Total": total,
                "Avaliados": avaliados, "Conforme": conf, "Parcial": parc,
                "Não Conforme": nconf, "Não Avaliado": nav,
                "pct_av": (avaliados/total*100.0) if total else 0.0,
                "pct_conf": (conf/total*100.0) if total else 0.0,
                "pct_nc_avaliados": round((nconf / avaliados * 100.0) if avaliados else 0.0, 2),
                "% Avaliado": f"{(avaliados/total*100.0) if total else 0.0:.2f}",
                "% Conforme": f"{(conf/total*100.0) if total else 0.0:.2f}",
            })
        rows.sort(key=lambda d: natural_sort_key(d.get("Capítulo", "")))
        return rows

# ── AppState ──────────────────────────────────────────────────────────────────
class AppState:
    def __init__(self):
        self.xlsx_path = DEFAULT_XLSX
        self.engine    = ExcelEngine(self.xlsx_path)
        self._save_lock = threading.Lock()
        self.pending_saves = 0
        self.last_save_at = ""
        self.last_save_error = ""
    def set_xlsx(self, path):
        self.xlsx_path = path
        try:
            self.engine.invalidate_read_cache()
        except Exception:
            pass
        self.engine    = ExcelEngine(self.xlsx_path)
    def reload(self):
        self.engine.invalidate_read_cache()
        self.engine.list_sheets()
    def begin_save(self):
        with self._save_lock:
            self.pending_saves += 1
            self.last_save_error = ""
    def end_save(self, ok: bool, error: str = ""):
        with self._save_lock:
            self.pending_saves = max(0, self.pending_saves - 1)
            if ok:
                self.last_save_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                self.last_save_error = ""
            else:
                self.last_save_error = (error or "Erro ao salvar").strip()
    def is_save_pending(self) -> bool:
        with self._save_lock:
            return self.pending_saves > 0

STATE = AppState()

def save_items_to_excel(xlsx_path, sheet_name, items, changed_by: str = ""):
    ExcelEngine(xlsx_path).save_items(sheet_name, items, changed_by=changed_by)


def start_background_save(sheet_name: str, items: List[RowItem], changed_by: str = ""):
    STATE.begin_save()

    def worker():
        try:
            save_items_to_excel(STATE.xlsx_path, sheet_name, items, changed_by=changed_by)
            STATE.reload()
            STATE.end_save(True)
        except Exception as ex:
            logging.exception('Falha ao salvar em background: %s', ex)
            STATE.end_save(False, str(ex))

    threading.Thread(target=worker, name='farol-save-worker', daemon=True).start()

async def handle_upload(e):
    safe_mkdir(DATA_DIR)
    try:    raw = await e.file.read()
    except Exception as ex: notify_err(f"Erro ao ler: {ex}"); return
    if not is_valid_xlsx_bytes(raw):
        notify_err("Arquivo inválido."); return
    with open(DEFAULT_XLSX, "wb") as f: f.write(raw)
    STATE.set_xlsx(DEFAULT_XLSX)
    try:    STATE.reload()
    except Exception as ex: notify_warn(f"Salvo, mas erro: {ex}"); return
    notify_ok("Arquivo carregado!")
    ui.navigate.to("/")

# ── Widget: barra dupla de progresso por seção ────────────────────────────────
def section_stats_widget(stats: List[SectionStat], max_show: int = 0):
    """Lista compacta de seções com barra dupla azul/verde."""
    if not stats: return
    show = stats if max_show == 0 else stats[:max_show]
    for s in show:
        with ui.element("div").classes("w-full py-1"):
            with ui.row().classes("w-full items-center justify-between gap-1"):
                ui.label(s.nome).classes("text-xs font-semibold flex-1").style(
                    "white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:60%")
                with ui.row().classes("items-center gap-2 text-xs opacity-70"):
                    ui.label(f"{s.avaliados}/{s.total}")
                    ui.label(f"{s.pct_av:.2f}%").classes("font-semibold text-sky-600")
                    ui.label(f"✔ {s.pct_conf:.2f}%").classes("font-semibold text-emerald-600")
            with ui.element("div").classes("relative w-full h-2 rounded-full bg-gray-200"):
                ui.element("div").classes("absolute top-0 left-0 h-2 rounded-full bg-sky-300").style(
                    f"width:{clamp01(s.pct_av/100)*100:.1f}%")
                ui.element("div").classes("absolute top-0 left-0 h-2 rounded-full bg-emerald-400").style(
                    f"width:{clamp01(s.pct_conf/100)*100:.1f}%")
    if max_show and len(stats) > max_show:
        ui.label(f"... e mais {len(stats)-max_show} seções").classes("text-xs opacity-50 mt-1")

# ── Widget: tabela de seções (dashboard) ─────────────────────────────────────
def section_table_widget(stats: List[SectionStat]):
    """
    Tabela completa de seções — usa APENAS ui.row/ui.label para evitar
    o bug 'Element has no attribute add' com tags HTML brutas no NiceGUI >= 3.
    """
    if not stats: return

    # Legenda
    with ui.row().classes("items-center gap-4 text-xs opacity-60 mb-1"):
        ui.element("div").classes("w-3 h-3 rounded-full bg-sky-300 inline-block")
        ui.label("Azul = % Avaliado")
        ui.element("div").classes("w-3 h-3 rounded-full bg-emerald-400 inline-block")
        ui.label("Verde = % Conforme")

    COLS = ["Seção", "Total", "Aval", "Conf", "Parcial", "NC", "NA", "% Aval", "% Conf", "Progresso"]
    COL_W = ["flex-1 min-w-[160px]", "w-12 text-center", "w-12 text-center",
             "w-14 text-center", "w-16 text-center", "w-10 text-center",
             "w-10 text-center", "w-16 text-center", "w-16 text-center", "w-24"]

    # Cabeçalho
    with ui.row().classes("w-full bg-gray-100 rounded-t border-b px-2 py-1 gap-1"):
        for col, w in zip(COLS, COL_W):
            ui.label(col).classes(f"text-xs font-semibold {w}")

    # Linhas
    for s in stats:
        with ui.row().classes("w-full border-b hover:bg-gray-50 px-2 py-1 gap-1 items-center"):
            ui.label(s.nome).classes(f"text-xs font-semibold {COL_W[0]}").style(
                "white-space:nowrap;overflow:hidden;text-overflow:ellipsis;")
            ui.label(str(s.total)).classes(f"text-xs {COL_W[1]}")
            ui.label(str(s.avaliados)).classes(f"text-xs {COL_W[2]}")
            ui.label(str(s.conforme)).classes(f"text-xs text-emerald-700 font-semibold {COL_W[3]}")
            ui.label(str(s.parcial)).classes(f"text-xs text-amber-700 {COL_W[4]}")
            ui.label(str(s.nao_conforme)).classes(f"text-xs text-red-600 {COL_W[5]}")
            ui.label(str(s.nav)).classes(f"text-xs text-gray-500 {COL_W[6]}")
            ui.label(f"{s.pct_av:.2f}%").classes(f"text-xs text-sky-600 font-semibold {COL_W[7]}")
            ui.label(f"{s.pct_conf:.2f}%").classes(f"text-xs text-emerald-600 font-semibold {COL_W[8]}")
            with ui.element("div").classes(f"relative h-2 rounded-full bg-gray-200 {COL_W[9]}"):
                ui.element("div").classes("absolute top-0 left-0 h-2 rounded-full bg-sky-300").style(
                    f"width:{clamp01(s.pct_av/100)*100:.1f}%")
                ui.element("div").classes("absolute top-0 left-0 h-2 rounded-full bg-emerald-400").style(
                    f"width:{clamp01(s.pct_conf/100)*100:.1f}%")

# ── Modal de edição ───────────────────────────────────────────────────────────
def open_edit_dialog(sheet_name, item: RowItem, on_saved):
    local_av = item.avaliacao   or "Não Avaliado"
    local_ev = item.evidencias  or ""
    local_pr = item.prazo       or ""
    audit_user = item.usuario_alteracao or "-"
    audit_date = item.data_alteracao or "-"
    
    with ui.dialog() as dialog:
        with ui.card().classes("farol-modal-card max-w-[980px] max-w-[96vw] p-0 shadow-2xl border border-slate-200 bg-white").style("""
                display: flex;
                flex-direction: column;
                width: 100%;
                position: relative;
                padding-bottom: calc(env(safe-area-inset-bottom, 0px) + 14px);
            """):
            # Botão fechar flutuante no topo direito
            ui.button(icon="close", on_click=dialog.close).props("flat round").style("""
                position: absolute;
                top: 0.75rem;
                right: 0.75rem;
                z-index: 10;
                color: #6b7280;
            """)
            
            # Info - Seção / Requisito Global / Requisito Específico (full width background)
            with ui.element("div").style("background: #f9fafb; border-bottom: 1px solid #e5e7eb; width: 100%; box-sizing: border-box"):
                with ui.column().classes("px-6 py-3 gap-2").style("padding-right: 3rem"):
                    # Seção em negrito com fonte de título
                    if item.secao.strip():
                        ui.label(item.secao).classes("font-bold").style("font-size: 1.1rem; color: #4f46e5")
                    
                    # Requisito Global
                    rg = (item.requisito_geral or "").strip() or "Sem título"
                    ui.label("Requisito Global").classes("text-xs uppercase opacity-60 mt-1").style("font-size: 0.7rem")
                    ui.label(rg).classes("text-sm").style("line-height: 1.4")
                    
                    # Requisito Específico (se existir)
                    if (item.requisito or "").strip():
                        ui.label("Requisito Específico").classes("text-xs uppercase opacity-60 mt-2").style("font-size: 0.7rem")
                        ui.label(item.requisito).classes("text-sm").style("line-height: 1.4")
                    with ui.card().classes("w-full bg-slate-50 border border-slate-200 mt-3"):
                        ui.label("Auditoria").classes("text-xs uppercase tracking-wide opacity-60")
                        with ui.row().classes("w-full items-center justify-between gap-3 text-sm"):
                            ui.label(f"Última alteração: {audit_date}").classes("font-medium")
                            ui.label(f"Alterado por: {audit_user}").classes("font-medium text-indigo-700")
            
            # Content area - com altura definida para scroll funcionar
            with ui.element("div").style("""
                display: flex;
                flex-direction: column;
                gap: 1rem;
                width: 100%;
                box-sizing: border-box;
                padding-bottom: 0.5rem;
            """):
                
                # Status e Prazo em linha - lado a lado, 100% largura
                with ui.element("div").style("""
                    display: flex;
                    gap: 1rem;
                    width: 100%;
                    padding: 1.5rem 1.5rem 0 1.5rem;
                    box-sizing: border-box;
                """):
                    # Status - 50%
                    with ui.card().style("""
                        border-radius: 1rem;
                        background: #f8fafc;
                        border: 1px solid #e2e8f0;
                        padding: 1rem;
                        flex: 1;
                        box-sizing: border-box;
                    """):
                        ui.label("Status").classes("font-semibold text-base mb-2")
                        status_select = ui.select(AVALIACAO_OPCOES, value=local_av
                                                  ).classes("w-full").props("outlined dense")
                    
                    # Prazo - 50%
                    with ui.card().style("""
                        border-radius: 1rem;
                        background: #f8fafc;
                        border: 1px solid #e2e8f0;
                        padding: 1rem;
                        flex: 1;
                        box-sizing: border-box;
                    """):
                        ui.label("Prazo").classes("font-semibold text-base mb-2")
                        # Converte para formato DD/MM/YYYY para exibição
                        display_pr = ""
                        if local_pr:
                            try:
                                from datetime import datetime
                                d = datetime.strptime(str(local_pr), "%Y-%m-%d")
                                display_pr = d.strftime("%d/%m/%Y")
                            except:
                                display_pr = local_pr
                        date_input = ui.input(value=display_pr, placeholder="DD/MM/YYYY"
                                             ).classes("w-full").props("outlined dense mask='##/##/####' fill-mask='_'")
                
                # Evidências - 100% da largura
                with ui.card().style("""
                    border-radius: 1rem;
                    background: #f8fafc;
                    border: 1px solid #e2e8f0;
                    padding: 1rem;
                    display: flex;
                    flex-direction: column;
                    margin: 0 1.5rem 1.5rem 1.5rem;
                    width: calc(100% - 3rem);
                    box-sizing: border-box;
                    height: clamp(180px, 28vh, 260px);
                    min-height: 180px;
                """):
                    ui.label("Evidências").classes("font-semibold text-base mb-1")
                    ui.label("Links, prints, protocolos ou qualquer evidência da conformidade").classes("text-xs opacity-70 mb-3")
                    evid = ui.textarea(value=local_ev).classes("farol-textarea w-full").props("outlined").style("height: clamp(120px, 20vh, 180px); overflow-y: auto")
            
            ui.separator()
            
            # Footer
            with ui.row().classes("w-full items-center justify-end gap-3 px-6").style("""
                padding-top: 0.75rem;
                padding-bottom: calc(env(safe-area-inset-bottom, 0px) + 18px);
                min-height: 72px;
                background: white;
                position: sticky;
                bottom: 0;
                z-index: 5;
            """):
                ui.button("Cancelar", icon="close", on_click=dialog.close).props("outline")
                
                async def do_save():
                    save_btn.enabled = False
                    save_btn.props("loading")
                    try:
                        from datetime import datetime
                        
                        item.avaliacao   = (status_select.value or "Não Avaliado").strip()
                        item.evidencias  = (evid.value or "").strip()
                        
                        prazo_str = (date_input.value or "").strip()
                        if prazo_str:
                            try:
                                d = datetime.strptime(prazo_str, "%d/%m/%Y")
                                hoje = datetime.now()
                                if d.date() < hoje.date():
                                    notify_err("❌ Data não pode ser no passado!")
                                    save_btn.props(remove="loading")
                                    save_btn.enabled = True
                                    return
                                item.prazo = d.strftime("%Y-%m-%d")
                            except ValueError:
                                notify_err("❌ Data inválida! Use o formato DD/MM/YYYY com valores reais")
                                save_btn.props(remove="loading")
                                save_btn.enabled = True
                                return
                        else:
                            item.prazo = ""
                        
                        username = current_username() or "sistema"
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        item.usuario_alteracao = username
                        item.data_alteracao = timestamp

                        if sheet_name in STATE.engine.cache_items:
                            for cached_item in STATE.engine.cache_items[sheet_name]:
                                if cached_item.excel_row == item.excel_row:
                                    cached_item.avaliacao = item.avaliacao
                                    cached_item.evidencias = item.evidencias
                                    cached_item.prazo = item.prazo
                                    cached_item.usuario_alteracao = username
                                    cached_item.data_alteracao = timestamp
                        
                        STATE.engine.cache_secstats.pop(sheet_name, None)
                        start_background_save(sheet_name, [item], username)
                        notify_ok("✏️ Alteração enviada para gravação. Aguarde o disquete verde antes de baixar.")
                        on_saved()
                        dialog.close()
                        
                    except Exception as ex:
                        notify_err(f"Erro ao preparar salvamento: {ex}")
                    finally:
                        save_btn.props(remove="loading")
                        save_btn.enabled = True
                
                save_btn = ui.button("Salvar", icon="save", on_click=do_save).props("color=primary")
    dialog.open()

def item_row_compact(sheet_name, item: RowItem, on_saved):
    with ui.row().classes("w-full items-center justify-between gap-3 py-2 border-b"):
        ui.label(item.requisito).classes("flex-1 text-sm").style("white-space:normal;line-height:1.2;")
        with ui.row().classes("items-center gap-2"):
            ui.badge(item.avaliacao, color=status_color(item.avaliacao))
            ui.button("Avaliar", icon="edit",
                      on_click=lambda: open_edit_dialog(sheet_name, item, on_saved)
                      ).props("outline size=sm")

def render_items_grouped(sheet_name, items: List[RowItem], on_saved):
    has_secao = any(it.secao.strip() for it in items)
    if not has_secao:
        groups: Dict[str, List[RowItem]] = {}
        for it in items:
            groups.setdefault(it.requisito_geral or "(Sem título)", []).append(it)
        for rg, its in groups.items():
            with ui.card().classes("w-full"):
                ui.label(rg).classes("text-base font-bold bg-gray-100 p-2 rounded")
                for it in its: item_row_compact(sheet_name, it, on_saved)
    else:
        top: Dict[str, Dict[str, List[RowItem]]] = {}
        for it in items:
            sec = it.secao.strip() or "(Sem seção)"
            rg  = it.requisito_geral.strip() or "(Sem título)"
            top.setdefault(sec, {}).setdefault(rg, []).append(it)
        for sec, sub in top.items():
            with ui.card().classes("w-full border-l-4 border-indigo-400 bg-indigo-50"):
                ui.label(sec).classes("text-base font-extrabold text-indigo-800 px-2 py-1")
                for rg, its in sub.items():
                    with ui.card().classes("w-full mt-2"):
                        ui.label(rg).classes("text-sm font-bold bg-gray-100 p-2 rounded text-gray-700")
                        for it in its: item_row_compact(sheet_name, it, on_saved)

def apply_filters(items: List[RowItem], search: str, status: str) -> List[RowItem]:
    s = norm(search); st = status.strip()
    return [it for it in items
            if (st == "Todos" or it.avaliacao == st)
            and (not s or s in norm(it.requisito) or s in norm(it.requisito_geral) or s in norm(it.secao))]


def create_loading_screen(title: str, subtitle: str):
    with ui.column().classes("w-full max-w-[1400px] mx-auto p-4 gap-4 min-h-[70vh] justify-center") as root:
        with ui.row().classes('w-full justify-center items-center flex-1'):
            with ui.card().classes("w-full max-w-3xl rounded-3xl shadow-xl border-2 border-emerald-200 bg-gradient-to-br from-emerald-700 via-emerald-600 to-sky-700 text-white overflow-hidden") as card:
                with ui.column().classes("w-full items-center justify-center gap-4 p-8 md:p-12"):
                    ui.icon('rocket_launch').classes('text-5xl text-white animate-pulse')
                    ui.label(title).classes('text-2xl md:text-3xl font-bold text-center')
                    ui.label(subtitle).classes('text-sm md:text-base text-center opacity-90 max-w-2xl')
                    progress = ui.linear_progress(value=0.05).classes('w-full max-w-2xl')
                    progress.props('rounded size=16px color=light-green-3 track-color=cyan-10')
                    status = ui.label('Preparando cenário...').classes('text-sm opacity-95 font-medium')
                    hint = ui.label('Carregando dados do Excel, estatísticas e capítulos.').classes('text-xs opacity-80 text-center')
        content = ui.column().classes('w-full gap-4')
    return card, progress, status, hint, content


def render_home_content(content, chapters, dash_rows):
    total = sum(r["Total"] for r in dash_rows) if dash_rows else 0
    avaliados = sum(r["Avaliados"] for r in dash_rows) if dash_rows else 0
    conforme = sum(r["Conforme"] for r in dash_rows) if dash_rows else 0
    pct_av = (avaliados / total * 100.0) if total else 0.0
    pct_conf = (conforme / total * 100.0) if total else 0.0
    metrics = {r["Capítulo"]: r for r in dash_rows}
    search = {"v": ""}

    with content:
        if not chapters:
            ui.label("Nenhum capítulo visível encontrado.").classes("opacity-70")
            ui.label("Abra Configurações para importar o arquivo Excel.").classes("text-sm opacity-70")
            return
        with ui.element("div").classes("grid grid-cols-1 lg:grid-cols-3 gap-4 w-full"):
            with ui.card().classes("w-full rounded-2xl shadow-sm border lg:col-span-2"):
                with ui.row().classes("w-full items-start justify-between"):
                    with ui.column().classes("gap-1"):
                        ui.label("Visão Geral").classes("text-lg font-bold")
                        ui.label(f"Total: {total} • Avaliados: {avaliados} • Conforme: {conforme}").classes("text-sm opacity-70")
                    ui.icon("insights").classes("opacity-50")
                ui.separator()
                with ui.element("div").classes("grid grid-cols-1 sm:grid-cols-2 gap-4 w-full"):
                    with ui.card().classes("w-full rounded-2xl border bg-sky-50 border-sky-200"):
                        ui.label("% Avaliados").classes("text-xs uppercase tracking-wide opacity-70")
                        ui.label(f"{pct_av:.2f}%").classes("text-2xl font-bold")
                        ui.linear_progress(value=round(clamp01(pct_av/100), 2)).props("show-value=false rounded size=10px")
                    with ui.card().classes("w-full rounded-2xl border bg-emerald-50 border-emerald-200"):
                        ui.label("% Conforme").classes("text-xs uppercase tracking-wide opacity-70")
                        ui.label(f"{pct_conf:.2f}%").classes("text-2xl font-bold")
                        ui.linear_progress(value=round(clamp01(pct_conf/100), 2)).props("show-value=false rounded size=10px")
            with ui.card().classes("w-full rounded-2xl shadow-sm border"):
                ui.label("Buscar capítulo").classes("text-lg font-bold")
                inp = ui.input(label="Filtro", placeholder="Ex.: Gestão, Segurança ...").classes("w-full").props("outlined clearable")
                ui.separator()
                with ui.row().classes("w-full items-center justify-between"):
                    ui.badge(f"Capítulos: {len(chapters)}", color="grey")
                    ui.badge(f"Visíveis: {len(chapters)}", color="grey")
        ui.label("Capítulos").classes("text-lg font-semibold")

        @ui.refreshable
        def chapters_grid():
            q = norm(search["v"])
            filtered = [si for si in chapters if not q or q in norm(si.name)]
            if not filtered:
                ui.label("Nenhum capítulo encontrado.").classes("opacity-70")
                return
            with ui.element("div").classes("grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 xl:grid-cols-4 gap-4 w-full"):
                for si in filtered:
                    r = metrics.get(si.name) or {}
                    pct_av_c = r.get("pct_av", 0.0)
                    pct_cf_c = r.get("pct_conf", 0.0)
                    av_c = r.get("Avaliados", 0)
                    tot_c = r.get("Total", 0)
                    with ui.card().classes("w-full rounded-2xl shadow-sm border hover:shadow"):
                        with ui.row().classes("w-full items-start justify-between gap-2 cursor-pointer").on(
                            "click", lambda slug=si.slug: ui.navigate.to(f"/capitulo/{slug}")):
                            ui.label(si.name).classes("font-bold text-base flex-1")
                            ui.icon("chevron_right").classes("opacity-40")
                        ui.label(f"Avaliados: {av_c}/{tot_c}").classes("text-sm opacity-70")
                        with ui.element("div").classes("w-full py-2"):
                            with ui.row().classes("w-full items-center justify-between gap-1"):
                                ui.label(si.name).classes("text-xs font-semibold flex-1").style(
                                    "white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:60%")
                                with ui.row().classes("items-center gap-2 text-xs opacity-70"):
                                    ui.label(f"{av_c}/{tot_c}")
                                    ui.label(f"{pct_av_c:.2f}%").classes("font-semibold text-sky-600")
                                    ui.label(f"✔ {pct_cf_c:.2f}%").classes("font-semibold text-emerald-600")
                            with ui.element("div").classes("relative w-full h-2 rounded-full bg-gray-200"):
                                ui.element("div").classes("absolute top-0 left-0 h-2 rounded-full bg-sky-300").style(
                                    f"width:{clamp01(pct_av_c/100)*100:.1f}%")
                                ui.element("div").classes("absolute top-0 left-0 h-2 rounded-full bg-emerald-400").style(
                                    f"width:{clamp01(pct_cf_c/100)*100:.1f}%")

        def on_change_search(e):
            search["v"] = e.value or ""
            chapters_grid.refresh()

        inp.on("update:model-value", on_change_search)
        chapters_grid()


def render_dashboard_content(content, rows):
    with content:
        ui.label("Dashboard").classes("text-2xl font-bold")
        if not rows:
            ui.label("Sem dados. Verifique as abas.").classes("opacity-70")
            return

        total = sum(r["Total"] for r in rows)
        conf = sum(r["Conforme"] for r in rows)
        parc = sum(r["Parcial"] for r in rows)
        nconf = sum(r["Não Conforme"] for r in rows)
        nav = sum(r["Não Avaliado"] for r in rows)
        avaliados = sum(r["Avaliados"] for r in rows)
        pct_conf = (conf / total * 100.0) if total else 0.0
        pct_av = (avaliados / total * 100.0) if total else 0.0

        def kpi(title, value, subtitle="", icon="insights", accent="border-slate-300", bg="bg-white"):
            with ui.card().classes(f"w-full rounded-2xl shadow-sm border {accent} {bg} min-h-[96px]"):
                with ui.row().classes("w-full items-start justify-between"):
                    with ui.column().classes("gap-1"):
                        ui.label(title).classes("text-xs uppercase tracking-wide opacity-70")
                        ui.label(value).classes("text-2xl font-bold leading-tight")
                        if subtitle:
                            ui.label(subtitle).classes("text-xs opacity-60")
                    ui.icon(icon).classes("opacity-50")

        with ui.element("div").classes("grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-6 gap-4 w-full"):
            kpi("Total", str(total), "", "list_alt", "border-slate-200", "bg-slate-50")
            kpi("Avaliados", str(avaliados), f"{pct_av:.2f}% do total", "task_alt", "border-sky-200", "bg-sky-50")
            kpi("Conforme", str(conf), f"{pct_conf:.2f}% total", "check_circle", "border-emerald-200", "bg-emerald-50")
            kpi("Parcial", str(parc), "", "tune", "border-amber-200", "bg-amber-50")
            kpi("Não Conforme", str(nconf), "", "report_problem", "border-rose-200", "bg-rose-50")
            kpi("Não Avaliado", str(nav), "", "hourglass_empty", "border-slate-200", "bg-slate-50")

        with ui.card().classes("w-full"):
            ui.label("Progresso geral").classes("font-semibold")
            ui.label(f"% Avaliado: {pct_av:.2f}%").classes("text-sm opacity-70")
            ui.linear_progress(value=round(clamp01(pct_av/100), 2)).props("show-value=false rounded size=10px")
            ui.separator()
            ui.label(f"% Conforme: {pct_conf:.2f}%").classes("text-sm opacity-70")
            ui.linear_progress(value=round(clamp01(pct_conf/100), 2)).props("show-value=false rounded size=10px")

        color_map = {"Conforme":"#22c55e","Parcial":"#f59e0b","Não Conforme":"#ef4444","Não Avaliado":"#94a3b8"}
        with ui.element("div").classes("grid grid-cols-1 lg:grid-cols-2 gap-4 w-full"):
            pie_opt = {
                "tooltip":{"trigger":"item"}, "legend":{"top":"bottom"},
                "series":[{"name":"Status","type":"pie","radius":["35%","70%"],
                    "avoidLabelOverlap":True,
                    "itemStyle":{"borderRadius":6,"borderColor":"#fff","borderWidth":2},
                    "label":{"show":True,"formatter":"{b}: {c} ({d}%)"},
                    "data":[
                        {"value":conf,  "name":"Conforme",     "itemStyle":{"color":color_map["Conforme"]}},
                        {"value":parc,  "name":"Parcial",      "itemStyle":{"color":color_map["Parcial"]}},
                        {"value":nconf, "name":"Não Conforme", "itemStyle":{"color":color_map["Não Conforme"]}},
                        {"value":nav,   "name":"Não Avaliado", "itemStyle":{"color":color_map["Não Avaliado"]}},
                    ]}],
            }
            with ui.card().classes("w-full rounded-2xl shadow-sm border"):
                ui.label("Distribuição de status").classes("font-semibold")
                ui.echart(pie_opt).classes("w-full h-[340px]")

            top10 = sorted(rows, key=lambda r: r.get("pct_nc_avaliados", 0), reverse=True)[:10]
            bar_opt = {
                "tooltip": {
                    "trigger": "axis",
                    "axisPointer": {"type": "shadow"},
                    "formatter": "{b}<br/>Não Conforme: {c}%"
                },
                "grid": {"left": "3%", "right": "3%", "bottom": "0%", "containLabel": True},
                "xAxis": {
                    "type": "value",
                    "name": "% Não Conforme",
                    "axisLabel": {"formatter": "{value}%"}
                },
                "yAxis": {"type": "category", "data": [r["Capítulo"] for r in top10]},
                "series": [{
                    "name": "% Não Conforme",
                    "type": "bar",
                    "data": [r.get("pct_nc_avaliados", 0) for r in top10],
                    "itemStyle": {"color": color_map["Não Conforme"]},
                    "label": {"show": True, "position": "right", "formatter": "{c}%"}
                }],
            }
            with ui.card().classes("w-full rounded-2xl shadow-sm border"):
                ui.label("Top 10 (mais não conformes)").classes("font-semibold")
                ui.label("Percentual de 'Não Conforme' em relação aos itens avaliados (exclui Não Avaliado)").classes("text-xs opacity-70 mb-2")
                ui.echart(bar_opt).classes("w-full h-[340px]")

        with ui.card().classes("w-full"):
            with ui.row().classes("w-full items-center justify-between"):
                ui.label("Resumo por Capítulo e Seção").classes("font-semibold text-base")
                def export_dashboard_csv():
                    cols = ["Capítulo","Total","Avaliados","Conforme","Parcial",
                            "Não Conforme","Não Avaliado","% Avaliado","% Conforme"]
                    buf = io.StringIO()
                    w = csv.DictWriter(buf, fieldnames=cols, delimiter=";")
                    w.writeheader()
                    for r in rows:
                        w.writerow({c: r.get(c,"") for c in cols})
                    ui.download(buf.getvalue().encode("utf-8-sig"), "dashboard_resumo.csv")
                ui.button("Exportar CSV", icon="download", on_click=export_dashboard_csv).props("outline")

            for r in rows:
                cap = r["Capítulo"]
                slug = r.get("slug", slugify(cap))
                try:
                    sec_stats = STATE.engine.section_stats(cap)
                except Exception:
                    sec_stats = []

                with ui.expansion(cap, icon="folder").classes("w-full border rounded-lg mb-1"):
                    with ui.row().classes("w-full items-center gap-2 flex-wrap py-1"):
                        ui.badge(f"Total: {r['Total']}", color="grey")
                        ui.badge(f"Aval: {r['Avaliados']}", color="blue")
                        ui.badge(f"Conf: {r['Conforme']}", color="green")
                        ui.badge(f"Parcial: {r['Parcial']}", color="orange")
                        ui.badge(f"NC: {r['Não Conforme']}", color="red")
                        ui.badge(f"NA: {r['Não Avaliado']}", color="grey")
                        ui.badge(f"% Aval: {r['% Avaliado']}%", color="blue")
                        ui.badge(f"% Conf: {r['% Conforme']}%", color="green")
                        ui.button("Abrir capítulo", icon="open_in_new", on_click=lambda s=slug: ui.navigate.to(f"/capitulo/{s}")).props("flat dense size=sm")
                    if sec_stats:
                        ui.label("Seções internas:").classes("text-xs font-semibold uppercase opacity-60 mt-2 mb-1")
                        section_table_widget(sec_stats)

# ── Pages ─────────────────────────────────────────────────────────────────────
@ui.page("/")
def page_home():
    if not require_login():
        return
    render_navbar("home")
    ui.page_title("Farol ONA - Home")
    loading_card, progress, status, hint, content = create_loading_screen(
        "Carregando base ONA",
        "Montando a home, indicadores e capítulos. Isso pode levar alguns segundos em planilhas maiores.",
    )

    async def load_home_data():
        try:
            progress.value = 0.12
            status.text = "Autenticando acesso ao painel..."
            hint.text = "Separando sessão, cache e contexto do usuário."
            await asyncio.sleep(0.05)

            progress.value = 0.32
            status.text = "Lendo capítulos do Excel..."
            hint.text = "Detectando abas visíveis e preparando a ordem correta."
            if not STATE.engine.sheet_infos:
                await asyncio.to_thread(STATE.reload)
            chapters = await asyncio.to_thread(STATE.engine.visible_chapters)

            progress.value = 0.68
            status.text = "Calculando indicadores gerais..."
            hint.text = "Resumo da home, porcentagens e desempenho por capítulo."
            dash_rows = await asyncio.to_thread(STATE.engine.compute_dashboard)

            progress.value = 0.92
            status.text = "Renderizando interface..."
            hint.text = "Quase lá. Aplicando os dados na tela."
            await asyncio.sleep(0.05)

            loading_card.delete()
            render_home_content(content, chapters, dash_rows)
        except Exception as ex:
            progress.value = 1.0
            status.text = "Falha ao carregar a home."
            hint.text = str(ex)
            with content:
                ui.label(f"Não foi possível carregar a Home: {ex}").classes("text-red-600")

    ui.timer(0.05, load_home_data, once=True)


@ui.page("/dashboard")
def page_dashboard():
    if not require_login():
        return
    render_navbar("dashboard")
    ui.page_title("Dashboard - Farol ONA")
    loading_card, progress, status, hint, content = create_loading_screen(
        "Carregando dashboard",
        "Preparando gráficos, KPIs e estatísticas por capítulo. Em arquivos maiores isso demora um pouco.",
    )

    async def load_dashboard_data():
        try:
            progress.value = 0.18
            status.text = "Lendo estrutura da planilha..."
            hint.text = "Validando abas e atualizando o cache de leitura."
            if not STATE.engine.sheet_infos:
                await asyncio.to_thread(STATE.reload)

            progress.value = 0.55
            status.text = "Calculando KPIs e distribuição de status..."
            hint.text = "Somando totais e gerando os insumos dos gráficos."
            rows = await asyncio.to_thread(STATE.engine.compute_dashboard)

            progress.value = 0.88
            status.text = "Montando os componentes do dashboard..."
            hint.text = "Aplicando gráficos, tabela e expansões por seção."
            await asyncio.sleep(0.05)

            loading_card.delete()
            render_dashboard_content(content, rows)
        except Exception as ex:
            progress.value = 1.0
            status.text = "Falha ao carregar o dashboard."
            hint.text = str(ex)
            with content:
                ui.label(f"Não foi possível carregar o Dashboard: {ex}").classes("text-red-600")

    ui.timer(0.05, load_dashboard_data, once=True)


@ui.page("/capitulo/{cap_slug}")
def page_capitulo(cap_slug: str):
    if not require_login():
        return
    render_navbar("home")
    ui.page_title("Capítulo - Farol ONA")
    if not STATE.engine.sheet_infos:
        try: STATE.reload()
        except: pass
    si = {s.slug: s for s in STATE.engine.visible_chapters()}.get(cap_slug)
    with ui.column().classes("w-full max-w-[1400px] mx-auto p-4 gap-3"):
        if not si:
            ui.label("Capítulo não encontrado.").classes("opacity-70"); return
        ui.label(si.name).classes("text-2xl font-bold")
        try:    items = STATE.engine.load_items(si.name)
        except Exception as ex:
            notify_err(f"Erro lendo '{si.name}': {ex}"); items = []
        if not items:
            ui.label("Nenhum item encontrado nesta aba.").classes("opacity-70"); return
        has_secao = any(it.secao.strip() for it in items)
        try:    sec_stats = STATE.engine.section_stats(si.name)
        except: sec_stats = []

        # Painel de seções no topo
        if sec_stats:
            label_secao = "Seções internas" if has_secao else "Requisitos gerais"
            with ui.card().classes("w-full"):
                with ui.row().classes("w-full items-center justify-between mb-2"):
                    ui.label(f"Progresso por {label_secao}").classes("font-semibold")
                    ui.badge(f"{len(sec_stats)} {'seções' if has_secao else 'req. gerais'}", color="indigo")
                with ui.row().classes("items-center gap-4 text-xs opacity-60 mb-3"):
                    ui.element("div").classes("w-3 h-3 rounded-full bg-sky-300 inline-block")
                    ui.label("Azul = % Avaliado")
                    ui.element("div").classes("w-3 h-3 rounded-full bg-emerald-400 inline-block")
                    ui.label("Verde = % Conforme")
                for s in sec_stats:
                    with ui.element("div").classes("w-full mb-3"):
                        with ui.row().classes("w-full items-center justify-between gap-2"):
                            ui.label(s.nome).classes("text-sm font-semibold flex-1").style(
                                "white-space:nowrap;overflow:hidden;text-overflow:ellipsis;")
                            with ui.row().classes("items-center gap-3 text-sm shrink-0"):
                                ui.label(f"{s.avaliados}/{s.total}").classes("opacity-60")
                                ui.badge(f"Aval: {s.pct_av:.2f}%",  color="blue")
                                ui.badge(f"Conf: {s.pct_conf:.2f}%", color="green")
                                if s.nao_conforme: ui.badge(f"NC: {s.nao_conforme}", color="red")
                                if s.nav:          ui.badge(f"NA: {s.nav}",          color="grey")
                        with ui.element("div").classes("relative w-full h-3 rounded-full bg-gray-200 mt-1"):
                            ui.element("div").classes("absolute top-0 left-0 h-3 rounded-full bg-sky-300").style(
                                f"width:{clamp01(s.pct_av/100)*100:.1f}%")
                            ui.element("div").classes("absolute top-0 left-0 h-3 rounded-full bg-emerald-400").style(
                                f"width:{clamp01(s.pct_conf/100)*100:.1f}%")

        # Filtros e lista
        search_value = {"v": ""}; status_value = {"v": "Todos"}; render_limit = {"v": INITIAL_RENDER_LIMIT}
        with ui.card().classes("w-full"):
            ui.label("Filtros").classes("font-semibold")
            with ui.row().classes("w-full items-center gap-3 flex-wrap"):
                inp = ui.input(label="Buscar requisito", placeholder="Digite para filtrar...").classes("w-96")
                sel = ui.select(["Todos"] + AVALIACAO_OPCOES, value="Todos", label="Status").classes("w-72")
                def do_export():
                    filtered = apply_filters(items, search_value["v"], status_value["v"])
                    data, fname = csv_bytes([x.to_dict() for x in filtered], f"{si.slug}.csv")
                    ui.download(data, fname)
                ui.button("Exportar CSV", icon="download", on_click=do_export).props("outline")

        def on_change_search(e):
            search_value["v"] = e.value or ""; render_limit["v"] = INITIAL_RENDER_LIMIT; list_block.refresh()
        def on_change_status(e):
            status_value["v"] = e.value or "Todos"; render_limit["v"] = INITIAL_RENDER_LIMIT; list_block.refresh()
        inp.on("update:model-value", on_change_search)
        sel.on("update:model-value", on_change_status)

        @ui.refreshable
        def list_block():
            filtered = apply_filters(items, search_value["v"], status_value["v"])
            conf_f  = sum(1 for x in filtered if x.avaliacao == "Conforme")
            parc_f  = sum(1 for x in filtered if x.avaliacao == "Parcialmente Conforme")
            nconf_f = sum(1 for x in filtered if x.avaliacao == "Não Conforme")
            nav_f   = sum(1 for x in filtered if x.avaliacao == "Não Avaliado")
            with ui.row().classes("w-full gap-2"):
                ui.badge(f"Total: {len(filtered)}",  color="grey")
                ui.badge(f"Conforme: {conf_f}",      color="green")
                ui.badge(f"Parcial: {parc_f}",       color="orange")
                ui.badge(f"Não Conforme: {nconf_f}", color="red")
                ui.badge(f"Não Avaliado: {nav_f}",   color="grey")
            limited = filtered[:render_limit["v"]]
            render_items_grouped(si.name, limited, on_saved=list_block.refresh)
            if len(filtered) > len(limited):
                with ui.row().classes("w-full justify-center mt-2"):
                    def load_more():
                        render_limit["v"] += RENDER_STEP; list_block.refresh()
                    ui.button(f"Carregar mais (+{RENDER_STEP})", icon="expand_more",
                              on_click=load_more).props("outline")
        list_block()


@ui.page("/login")
def page_login():
    ui.page_title("Login - Farol ONA")
    if get_session().get('authenticated'):
        ui.navigate.to('/')
        return
    with ui.column().classes("w-full max-w-md mx-auto min-h-screen justify-center p-6"):
        with ui.card().classes("w-full rounded-2xl shadow-lg border"):
            ui.label("Farol ONA").classes("text-2xl font-bold text-emerald-700")
            ui.label("Acesse com seu usuário e senha.").classes("text-sm opacity-70 mb-4")
            username_input = ui.input(label="Usuário").classes("w-full").props("outlined")
            password_input = ui.input(label="Senha", password=True, password_toggle_button=True).classes("w-full").props("outlined")

            async def do_login():
                user = verify_login(username_input.value or '', password_input.value or '')
                if not user:
                    notify_err('Usuário ou senha inválidos.')
                    return
                login_user(user)
                notify_ok('Login realizado com sucesso.')
                ui.navigate.to('/')

            ui.button("Entrar", icon="login", on_click=do_login).props("color=primary").classes("w-full mt-2")
            ui.label("No primeiro acesso, o sistema cria automaticamente o admin padrão.").classes("text-xs opacity-60 mt-3")


@ui.page("/config")
def page_config():
    if not require_admin():
        return
    render_navbar("config")
    ui.page_title("Configurações - Farol ONA")
    with ui.column().classes("w-full max-w-[1400px] mx-auto p-4 gap-4"):
        ui.label("Configurações").classes("text-2xl font-bold")
        with ui.card().classes("w-full"):
            ui.label("Arquivo Excel").classes("font-semibold")
            ui.label(f"Arquivo atual: {os.path.basename(STATE.engine.xlsx_path)}").classes("text-sm opacity-70")
            ui.label(f"Caminho: {STATE.engine.xlsx_path}").classes("text-xs opacity-60")
            def _reload():
                try:    STATE.reload(); notify_ok("Recarregado.")
                except Exception as ex: notify_err(f"Erro: {ex}")
            async def _handle_upload(e):
                safe_mkdir(DATA_DIR)
                try:    raw = await e.file.read()
                except Exception as ex: notify_err(f"Erro ao ler: {ex}"); return
                if not is_valid_xlsx_bytes(raw):
                    notify_err("Arquivo inválido."); return
                with open(DEFAULT_XLSX, "wb") as f: f.write(raw)
                STATE.set_xlsx(DEFAULT_XLSX); _reload()
            with ui.row().classes("items-center gap-3 mt-2"):
                ui.upload(label="Importar .xlsx", auto_upload=True, on_upload=_handle_upload).props("accept=.xlsx")
                ui.button("Recarregar", icon="refresh", on_click=_reload).props("outline")
            def _download_current():
                if STATE.is_save_pending():
                    notify_warn("Há salvamentos em andamento. Aguarde o disquete verde antes de baixar.")
                    return
                if STATE.last_save_error:
                    notify_err(f"Último salvamento falhou: {STATE.last_save_error}")
                    return
                if not os.path.exists(STATE.engine.xlsx_path):
                    notify_warn("Arquivo não encontrado."); return
                ui.download(open(STATE.engine.xlsx_path,"rb").read(), os.path.basename(STATE.engine.xlsx_path))
            def _download_backup():
                if STATE.is_save_pending():
                    notify_warn("Há salvamentos em andamento. Aguarde o disquete verde antes de baixar.")
                    return
                bak = STATE.engine.xlsx_path + ".bak"
                if not os.path.exists(bak):
                    notify_warn("Nenhum backup .bak ainda."); return
                ui.download(open(bak,"rb").read(), os.path.basename(bak))
            with ui.row().classes("gap-2 mt-2"):
                btn_download_current = ui.button("Baixar Excel atual", icon="download", on_click=_download_current).props("outline")
                btn_download_backup = ui.button("Baixar backup .bak", icon="history",  on_click=_download_backup).props("outline")
            save_status_hint = ui.label("").classes("text-xs mt-1")
            def refresh_download_state():
                pending = STATE.is_save_pending()
                btn_download_current.enabled = not pending
                btn_download_backup.enabled = not pending
                if pending:
                    save_status_hint.set_text("💾✏️ Salvando em background... o download fica bloqueado até concluir.")
                    save_status_hint.classes(replace='text-xs mt-1 text-amber-700')
                elif STATE.last_save_error:
                    save_status_hint.set_text(f"💾⚠️ Último salvamento falhou: {STATE.last_save_error}")
                    save_status_hint.classes(replace='text-xs mt-1 text-red-700')
                elif STATE.last_save_at:
                    save_status_hint.set_text(f"💾✅ Tudo salvo. Última confirmação: {STATE.last_save_at}.")
                    save_status_hint.classes(replace='text-xs mt-1 text-emerald-700')
                else:
                    save_status_hint.set_text("Backup .bak criado automaticamente a cada salvar concluído.")
                    save_status_hint.classes(replace='text-xs mt-1 opacity-60')
            refresh_download_state()
            ui.timer(0.7, refresh_download_state)
        with ui.card().classes("w-full"):
            ui.label("Usuários").classes("font-semibold")
            ui.label("Somente administradores podem criar novos acessos.").classes("text-sm opacity-70")
            with ui.row().classes("w-full items-end gap-3 flex-wrap mt-2"):
                new_user = ui.input(label="Novo usuário").classes("w-64").props("outlined")
                new_pass = ui.input(label="Senha", password=True, password_toggle_button=True).classes("w-64").props("outlined")
                admin_toggle = ui.checkbox("Administrador")

                async def handle_create_user():
                    ok, msg = create_user(new_user.value or '', new_pass.value or '', admin_toggle.value or False)
                    if ok:
                        notify_ok(msg)
                        new_user.value = ''
                        new_pass.value = ''
                        admin_toggle.value = False
                        users_table.refresh()
                    else:
                        notify_err(msg)

                ui.button("Criar usuário", icon="person_add", on_click=handle_create_user).props("color=primary")

            @ui.refreshable
            def users_table():
                data = load_users().get('users', [])
                if not data:
                    ui.label('Nenhum usuário cadastrado.').classes('opacity-70')
                    return
                with ui.column().classes('w-full gap-1 mt-4'):
                    for user in sorted(data, key=lambda u: (u.get('username') or '').lower()):
                        with ui.row().classes('w-full items-center justify-between border rounded px-3 py-2'):
                            with ui.column().classes('gap-0'):
                                ui.label(user.get('username') or '-').classes('font-semibold')
                                ui.label(user.get('created_at') or '-').classes('text-xs opacity-60')
                            with ui.row().classes('items-center gap-2'):
                                ui.badge('Admin' if user.get('is_admin') else 'Usuário', color='deep-orange' if user.get('is_admin') else 'grey')
            users_table()
        with ui.card().classes("w-full"):
            ui.label("Formatos suportados").classes("font-semibold")
            with ui.element("div").classes("grid grid-cols-1 md:grid-cols-2 gap-3"):
                with ui.card().classes("bg-gray-50 border"):
                    ui.label("Formato 2025").classes("font-semibold text-sm")
                    ui.label("Colunas: Requisito geral | Requisitos | Avaliação | Evidências | Observações | Data Alteração | Usuário Alteração").classes("text-xs opacity-70")
                    ui.label("Hierarquia: 2 níveis | Seções = Requisito geral").classes("text-xs opacity-70")
                with ui.card().classes("bg-indigo-50 border-indigo-200 border"):
                    ui.label("Formato 2026 ✨").classes("font-semibold text-sm text-indigo-700")
                    ui.label("Colunas: Seção | Requisito Global | Requisitos Específicos | Avaliação | Evidências | Observações | Data Alteração | Usuário Alteração").classes("text-xs opacity-70")
                    ui.label("Hierarquia: 3 níveis | Seções = coluna Seção").classes("text-xs opacity-70")
        with ui.card().classes("w-full"):
            ui.label("Diagnóstico").classes("font-semibold")
            xlsx_ok = is_valid_xlsx_path(STATE.engine.xlsx_path) if STATE.engine.exists() else False
            if not STATE.engine.exists():
                ui.label("⚠️ Arquivo não encontrado. Importe acima.").classes("text-red-500")
            elif not xlsx_ok:
                ui.label("⚠️ Arquivo inválido ou corrompido. Reimporte.").classes("text-red-500")
            elif not STATE.engine.sheet_infos:
                ui.label("Nenhuma aba detectada. Clique Recarregar.").classes("opacity-70")
            else:
                cap_vis = len(STATE.engine.visible_chapters())
                cap_tot = len(STATE.engine.sheet_infos)
                ui.label(f"✅ Válido — {cap_tot} abas ({cap_vis} visíveis)").classes("text-sm text-green-700 font-semibold")
                with ui.row().classes("w-full gap-2 flex-wrap"):
                    for si in STATE.engine.visible_chapters()[:15]:
                        ui.badge(si.name).classes("text-xs")
                    if cap_vis > 15:
                        ui.label(f"... e mais {cap_vis-15}").classes("text-sm opacity-60")

# ── Server ────────────────────────────────────────────────────────────────────

def main():
    # Inicializando o logger dentro do bloco para evitar o NameError
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
    logger = logging.getLogger(__name__)

    logger.info("Iniciando serviço Farol ONA...")
    safe_mkdir(DATA_DIR)
    ensure_users_file()
    STATE.set_xlsx(DEFAULT_XLSX)
    
    try:
        STATE.reload()
        logger.info("Planilha carregada com sucesso.")
    except Exception as ex:
        logger.error(f"Falha na inicialização da planilha: {ex}")

    # Configuração padronizada para produção com Proxy Reverso (Nginx)
    app_host = os.getenv("FAROL_HOST", "127.0.0.1")
    
    try:
        app_port = int(os.getenv("FAROL_PORT", 8100)) # Ajustado para 8100
    except ValueError:
        logger.warning("Porta inválida no ambiente, usando fallback 8100.")
        app_port = 8100 # Ajustado para 8100

    logger.info(f"NiceGUI configurado para escutar em {app_host}:{app_port}")
    
    try:
        ui.run(
            host=app_host, 
            port=app_port, 
            reload=False, 
            title="Farol ONA - Santa Casa", 
            favicon="🚀", 
            show=False,
            storage_secret=os.getenv("FAROL_STORAGE_SECRET", "farol-ona-storage-secret-troque-em-producao")
        )
    except Exception as e:
        logger.critical(f"Falha fatal ao iniciar a interface: {e}", exc_info=True)

if __name__ in {"__main__", "__mp_main__"}:
    main()
